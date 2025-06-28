import pandas as pd
import random
from datetime import datetime, timedelta
import os
import holidays
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# Feriados legales en Chile para el año deseado
feriados_chile = holidays.CL(years=2025)

## TODO: Completar
# revisar en el excel con el sarima --> 
def obtener_perfiles_para_dia(fecha: datetime) -> list:
    if fecha in feriados_chile or fecha.weekday() == 6:
        return []  # No hay perfiles en feriados o domingos
    perfiles = ["Perfil A", "Perfil B", "Perfil C", "Perfil D"]
    num_filas = random.randint(0, 9)
    return random.choices(perfiles, k=num_filas)

## TODO: Completar
## Aca hay una funcion de prediccion o no?
def calcular_ranking(perfil, tasa):
    return random.randint(1, 10)

## TODO: Completar
def calcular_venta_ponderada(perfil, ranking):
    return random.randint(50, 200)

def estilizar_excel(path_excel):
    wb = load_workbook(path_excel)
    ws = wb.active

    colores = {
        0: "B7E1FC",  # Fecha
        1: "C6EFCE",  # Tasa
        2: "FFF2CC",  # Ranking
        3: "E4C6EF",  # Venta
    }

    thin = Side(border_style="thin", color="000000")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, row in enumerate(ws.iter_rows(), start=1):
        for j, cell in enumerate(row, start=1):
            tipo = (j - 1) % 4
            fill = PatternFill(start_color=colores[tipo], end_color=colores[tipo], fill_type="solid")

            cell.fill = fill
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if i == 1:
                cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    wb.save(path_excel)
    print(f"🎨 Estilo aplicado a: {path_excel}")

def guardar_excel(df, archivo):
    try:
        df.to_excel(archivo, index=False)
        estilizar_excel(archivo)
    except PermissionError:
        print(f"❌ No se pudo guardar el archivo '{archivo}'. Asegúrese de que no esté abierto en Excel y vuelva a intentar.")
        input("📁 Cierre el archivo y presione Enter para volver a intentar...")
        guardar_excel(df, archivo)

def generar_excel(desde_fecha: str, cantidad_dias: int, archivo_salida: str):
    fecha_inicio = datetime.strptime(desde_fecha, "%d-%m-%Y")
    columnas_totales = []
    bloques_por_dia = []
    max_filas = 0

    for dia in range(cantidad_dias):
        fecha_actual = fecha_inicio + timedelta(days=dia)
        fecha_str = fecha_actual.strftime("%d-%m-%Y")
        columnas = [fecha_str, "Tasa", "Ranking", "Venta Ponderada"]
        columnas_totales.extend(columnas)

        perfiles_dia = obtener_perfiles_para_dia(fecha_actual)
        bloques_por_dia.append([[perfil, "", "", ""] for perfil in perfiles_dia])
        max_filas = max(max_filas, len(perfiles_dia))

    filas_finales = []
    for fila_idx in range(max_filas):
        fila = []
        for bloque in bloques_por_dia:
            if fila_idx < len(bloque):
                fila.extend(bloque[fila_idx])
            else:
                fila.extend(["", "", "", ""])
        filas_finales.append(fila)

    df = pd.DataFrame(filas_finales, columns=columnas_totales)
    guardar_excel(df, archivo_salida)
    print(f"✅ Excel generado correctamente como: {archivo_salida}")

def procesar_rankings(df: pd.DataFrame):
    columnas = df.columns.tolist()
    dias_procesados = 0
    dias_omitidos = []

    for i in range(0, len(columnas), 4):
        perfil_col = columnas[i]
        tasa_col = columnas[i + 1]
        ranking_col = columnas[i + 2]

        filas_validas = df[perfil_col].notna()
        tasas = df.loc[filas_validas, tasa_col]

        if filas_validas.sum() == 0:
            continue
        if tasas.isna().sum() > 0:
            print(f"⚠️ Día {perfil_col} tiene tasas incompletas. Se omite.")
            dias_omitidos.append(perfil_col)
            continue

        for fila in df[filas_validas].index:
            if pd.notna(df.at[fila, ranking_col]):
                continue
            perfil = df.at[fila, perfil_col]
            tasa = df.at[fila, tasa_col]
            df.at[fila, ranking_col] = calcular_ranking(perfil, float(tasa))

        dias_procesados += 1

    return df, dias_procesados, dias_omitidos

def procesar_ventas_ponderadas(df: pd.DataFrame):
    columnas = df.columns.tolist()
    dias_calculados = 0

    for i in range(0, len(columnas), 4):
        perfil_col = columnas[i]
        ranking_col = columnas[i + 2]
        venta_col = columnas[i + 3]

        filas_validas = df[perfil_col].notna() & df[ranking_col].notna()

        if not filas_validas.any():
            continue

        for fila in df[filas_validas].index:
            perfil = df.at[fila, perfil_col]
            ranking = df.at[fila, ranking_col]
            df.at[fila, venta_col] = calcular_venta_ponderada(perfil, ranking)

        dias_calculados += 1

    return df, dias_calculados

def menu_repeticion(df, archivo):
    while True:
        print("\n¿Desea repetir alguna etapa?")
        print("1. Rellenar tasas y recalcular rankings")
        print("2. Modificar rankings manualmente y recalcular ventas ponderadas")
        print("3. Salir")
        opcion = input("Ingrese una opción (1/2/3): ")

        if opcion == '1':
            input("✏️ Rellene las tasas en el Excel y presione Enter para continuar...")
            df = pd.read_excel(archivo)
            df, _, _ = procesar_rankings(df)
            guardar_excel(df, archivo)

        elif opcion == '2':
            input("✏️ Modifique manualmente los rankings y presione Enter para continuar...")
            df = pd.read_excel(archivo)
            df, _ = procesar_ventas_ponderadas(df)
            guardar_excel(df, archivo)

        elif opcion == '3':
            print("👋 Finalizando...")
            break
        else:
            print("❌ Opción no válida. Intente de nuevo.")
    

# Ejemplo de uso:
if __name__ == "__main__":
    archivo = "datos_clientes.xlsx"

    desde = input("📅 Ingrese la fecha de inicio (dd-mm-aaaa): ")
    while True:
        try:
            datetime.strptime(desde, "%d-%m-%Y")
            break
        except ValueError:
            desde = input("❌ Formato incorrecto. Use dd-mm-aaaa: ")

    while True:
        try:
            dias = int(input("📆 ¿Cuántos días desea simular?: "))
            break
        except ValueError:
            print("❌ Ingrese un número válido.")

    generar_excel(desde, dias, archivo)

    while True:
        input("\n✏️ Rellene las tasas en el Excel y presione Enter para continuar...")
        df = pd.read_excel(archivo)

        df, procesados, omitidos = procesar_rankings(df)
        guardar_excel(df, archivo)

        if omitidos:
            print(f"🔄 Días omitidos por tasas faltantes: {omitidos}")
            seguir = input("¿Desea rellenar las tasas faltantes y continuar? (s/n): ").lower()
            if seguir != 's':
                break
        else:
            break

    input("\n✏️ Puede modificar manualmente los rankings si desea. Presione Enter para continuar...")
    df = pd.read_excel(archivo)

    df, _ = procesar_ventas_ponderadas(df)
    guardar_excel(df, archivo)

    while True:
        vacios = []
        for i in range(0, len(df.columns), 4):
            perfil_col = df.columns[i]
            venta_col = df.columns[i+3]
            if df[perfil_col].notna().any() and df[venta_col].isna().any():
                vacios.append(perfil_col)

        if vacios:
            print(f"📌 Días con ventas ponderadas faltantes: {vacios}")
            seguir = input("¿Desea rellenar los rankings faltantes y continuar? (s/n): ").lower()
            if seguir == 's':
                df = pd.read_excel(archivo)
                df, _, _ = procesar_rankings(df)
                df, _ = procesar_ventas_ponderadas(df)
                guardar_excel(df, archivo)
            else:
                break
        else:
            break

    menu_repeticion(df, archivo)

    print("\n✅ Todo listo. Archivo final actualizado con ventas ponderadas.")
