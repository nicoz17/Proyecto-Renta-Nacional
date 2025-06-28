import pandas as pd
import random
from datetime import datetime, timedelta
import os
import holidays
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def generar_tasas_por_segmento():
    """Genera un diccionario de tasas predefinidas por segmento"""
    return {segmento: round(random.uniform(0.1, 1.0), 2) for segmento in range(1, 11)}

# Diccionario global de tasas por segmento
tasas_por_segmento = generar_tasas_por_segmento()

# Feriados legales en Chile para el año deseado
feriados_chile = holidays.CL(years=2025)

# Funciones (algunas de juguete, pueden ser reemplazadas luego)
def generar_clientes_dia(fecha: datetime) -> list:
    """Genera lista de clientes para un día (función de juguete)"""
    if fecha in feriados_chile or fecha.weekday() == 6:
        return []  # No hay clientes en feriados o domingos
    num_clientes = random.randint(0, 5)  # Entre 0 y 5 clientes por día
    return [f"Cliente-{i+1}" for i in range(num_clientes)]

def generar_cotizaciones_cliente(cliente: str) -> list:
    """Genera cotizaciones para un cliente (función de juguete)"""
    num_cotizaciones = random.randint(1, 3)  # Entre 1 y 3 cotizaciones por cliente
    return [f"Cotizacion-{i+1}" for i in range(num_cotizaciones)]

def asignar_tasa_automatica(segmento):
    """Asigna tasa automática basada en el segmento"""
    return tasas_por_segmento.get(segmento, 0.5)  # Default 0.5 si el segmento no existe

def rellenar_tasas_automaticamente(df):
    """Rellena todas las tasas vacías automáticamente basado en segmento"""
    columnas = df.columns.tolist()
    
    for i in range(0, len(columnas), 7):
        segmento_col = columnas[i+3]
        tasa_col = columnas[i+4]
        
        filas_con_segmento = df[segmento_col].notna()
        
        for fila in df[filas_con_segmento].index:
            if pd.isna(df.at[fila, tasa_col]):
                segmento = df.at[fila, segmento_col]
                df.at[fila, tasa_col] = asignar_tasa_automatica(segmento)
    
    return df

def menu_relleno_tasas(df, archivo):
    print("\n🔧 OPCIONES PARA RELLENAR TASAS")
    print("1. Rellenar automáticamente basado en segmentos")
    print("2. Rellenar manualmente en Excel")
    print("3. Ver tasas predefinidas por segmento")
    
    while True:
        opcion = input("Seleccione una opción (1/2/3): ")
        
        if opcion == '1':
            print("\n📊 Tasas predefinidas por segmento:")
            for seg, tasa in sorted(tasas_por_segmento.items()):
                print(f"Segmento {seg}: {tasa}")
            
            df = rellenar_tasas_automaticamente(df)
            guardar_excel(df, archivo)
            print("\n✅ Tasas asignadas automáticamente. Revise el archivo Excel.")
            
            # Opción para modificar manualmente después del automático
            if input("¿Desea modificar alguna tasa manualmente? (s/n): ").lower() == 's':
                input("✏️ Modifique las tasas en el Excel y presione Enter para continuar...")
                df = pd.read_excel(archivo)
                guardar_excel(df, archivo)
            
            return df
            
        elif opcion == '2':
            input("✏️ Abra el archivo Excel, rellene las tasas manualmente y presione Enter para continuar...")
            df = pd.read_excel(archivo)
            guardar_excel(df, archivo)
            return df
            
        elif opcion == '3':
            print("\n📊 Tasas predefinidas por segmento:")
            for seg, tasa in sorted(tasas_por_segmento.items()):
                print(f"Segmento {seg}: {tasa}")
        else:
            print("❌ Opción no válida. Intente de nuevo.")

def asignar_segmento(*args):  
    """Devuelve un segmento aleatorio entre 1 y 10."""  
    return random.randint(1, 10) 

def calcular_ranking(segmento: str, tasa: float) -> int:
    """Calcula ranking basado en segmento y tasa (función de juguete)"""
    return random.randint(1, 10)

def calcular_venta_ponderada(segmento, ranking):
    # Ejemplo básico (ajústalo según tu fórmula real)
    base = 100
    factor = 1.2
    return base * factor / ranking * segmento

def actualizar_ventas_y_resumen(path_excel):
    """Versión corregida que solo modifica ventas en Detalle y regenera Resumen"""
    try:
        # 1. Cargar el archivo manteniendo el formato
        wb = load_workbook(path_excel)
        
        if 'Detalle' not in wb.sheetnames:
            raise ValueError("No se encontró la hoja 'Detalle'")
            
        ws_detalle = wb['Detalle']
        
        # 2. Procesar solo las celdas necesarias en Detalle
        for row in ws_detalle.iter_rows(min_row=2):  # Saltar encabezados
            # Buscar columnas relevantes en cada fila (asumiendo estructura fija)
            for idx, cell in enumerate(row):
                # Identificar columnas por posición relativa (cada 7 columnas)
                if (idx % 7 == 3):  # Columna Segmento (índice 3 en cada bloque)
                    segmento = cell.value
                    segmento_col = cell.column
                elif (idx % 7 == 5):  # Columna Ranking (índice 5)
                    ranking = cell.value
                    ranking_col = cell.column
                elif (idx % 7 == 6):  # Columna Venta (índice 6)
                    venta_cell = cell
            
            # Solo procesar si tenemos todos los datos necesarios
            if (segmento is not None and 
                ranking is not None and 
                venta_cell is not None and 
                venta_cell.value is None):
                
                try:
                    if pd.notna(segmento) and pd.notna(ranking):
                        venta = calcular_venta_ponderada(int(segmento), int(ranking))
                        venta_cell.value = round(float(venta), 2)
                except Exception as e:
                    print(f"⚠️ Error en fila {row[0].row}: {str(e)}")
                    continue
        
        # 3. Guardar temporalmente los cambios en Detalle
        wb.save(path_excel)
        
        # 4. Actualizar hoja Resumen
        # Cargar datos actualizados de Detalle
        df_detalle = pd.read_excel(path_excel, sheet_name='Detalle')
        df_resumen = generar_hoja_resumen(df_detalle)
        
        # Guardar solo la hoja Resumen (preservando Detalle)
        with pd.ExcelWriter(path_excel, engine='openpyxl', mode='a') as writer:
            writer.book = load_workbook(path_excel)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            
            # Eliminar hoja Resumen existente
            if 'Resumen' in writer.book.sheetnames:
                writer.book.remove(writer.book['Resumen'])
            
            # Agregar nueva hoja Resumen
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        # 5. Reaplicar estilos
        estilizar_excel(path_excel)
        
        print("✅ Ventas asignadas y Resumen actualizado correctamente")
    
    except Exception as e:
        print(f"❌ Error crítico: {str(e)}")
        raise

def generar_hoja_resumen(df_detalle):
    """Versión corregida que maneja correctamente los bloques de columnas"""
    datos_resumen = []
    
    # Creamos una lista de columnas únicas temporalmente
    columnas_unicas = [f"Col_{i}" for i in range(len(df_detalle.columns))]
    df_temp = df_detalle.copy()
    df_temp.columns = columnas_unicas
    
    # Procesar cada bloque de 7 columnas (un día)
    for i in range(0, len(df_temp.columns), 7):
        if i + 7 > len(df_temp.columns):
            print(f"⚠️ Bloque incompleto en columnas {i} a {i+7}, omitiendo...")
            continue
            
        cols_dia = df_temp.columns[i:i+7]
        df_dia = df_temp[cols_dia].copy()
        
        # Verificar que tenemos exactamente 7 columnas
        if len(df_dia.columns) != 7:
            print(f"⚠️ Bloque en columnas {i} a {i+7} no tiene 7 columnas, tiene {len(df_dia.columns)}. Omitiendo...")
            continue
            
        # Renombrar columnas para procesamiento
        df_dia.columns = ['Fecha', 'Cliente', 'Cotizacion', 'Segmento', 'Tasa', 'Ranking', 'Venta']
        
        # Filtrar filas con clientes y segmentos válidos
        df_dia = df_dia[
            df_dia['Cliente'].notna() & 
            df_dia['Segmento'].notna() & 
            (df_dia['Segmento'] != '') & 
            (pd.to_numeric(df_dia['Segmento'], errors='coerce') > 0)
        ]
        
        if df_dia.empty:
            continue
            
        # Convertir columnas numéricas de forma segura
        df_dia['Tasa'] = pd.to_numeric(df_dia['Tasa'], errors='coerce')
        df_dia['Ranking'] = pd.to_numeric(df_dia['Ranking'], errors='coerce')
        df_dia['Venta'] = pd.to_numeric(df_dia['Venta'], errors='coerce')
        
        # Obtener la fecha real del primer elemento no nulo
        fecha = df_dia['Fecha'].iloc[0]
        
        # Agrupar por segmento (omitirá automáticamente segmentos vacíos/inválidos)
        for segmento, group in df_dia.groupby('Segmento'):
            # Verificación adicional de segmento válido
            if pd.isna(segmento) or segmento == '':
                continue
                
            try:
                segmento = int(segmento)
                if segmento <= 0:
                    continue
            except:
                continue
                
            num_cotizaciones = len(group)
            
            # Calcular tasa promedio con manejo seguro
            tasa_promedio = round(group['Tasa'].mean(), 2) if not group['Tasa'].isna().all() else 0
            
            # Distribución de rankings con manejo seguro
            ranking_counts = {}
            try:
                rankings = group['Ranking'].dropna()
                if not rankings.empty:
                    rankings = rankings.astype(int)
                    ranking_counts = rankings.value_counts().to_dict()
            except:
                ranking_counts = {}
            
            # Función segura para cálculo de porcentajes
            def calc_percent(rank):
                count = ranking_counts.get(rank, 0)
                return round((count / num_cotizaciones) * 100, 2) if num_cotizaciones > 0 else 0
            
            ranking_percent = {
                1: calc_percent(1),
                2: calc_percent(2),
                3: calc_percent(3),
                4: calc_percent(4),
                5: calc_percent(5),
                6: sum(calc_percent(r) for r in range(6, 11))  # Suma del 6 al 10
            }
            
            # Calcular venta ponderada con manejo seguro
            venta_ponderada = round(group['Venta'].sum(), 2) if not group['Venta'].isna().all() else 0
            
            # Agregar al resumen solo si hay datos válidos
            datos_resumen.append([
                fecha,
                segmento,
                num_cotizaciones,
                tasa_promedio,
                ranking_percent[1],
                ranking_percent[2],
                ranking_percent[3],
                ranking_percent[4],
                ranking_percent[5],
                ranking_percent[6],
                venta_ponderada
            ])
    
    # Crear DataFrame de resumen
    return pd.DataFrame(datos_resumen, columns=[
        "Fecha", "Segmento", "Numero de cotizaciones", "Tasa",
        "Ranking 1 %", "Ranking 2 %", "Ranking 3 %", "Ranking 4 %",
        "Ranking 5 %", "Ranking 6 y más %", "Venta ponderada"
    ])

def estilizar_excel(path_excel):
    """Aplica estilos a ambas hojas del Excel con el nuevo diseño"""
    wb = load_workbook(path_excel)
    
    if 'Resumen' in wb.sheetnames:
        ws_resumen = wb['Resumen']
        
        # Colores base para días alternados
        color_dia_par = "E6F3FF"  # Azul claro
        color_dia_impar = "E6FFE6"  # Verde claro
        color_fila_alterna_claro = "FFFFFF"  # Blanco
        color_fila_alterna_oscuro = "CCE5FF"  # Azul más oscuro para días azules
        color_fila_alterna_oscuro_verde = "CCFFCC"  # Verde más oscuro para días verdes
        
        # Aplicar estilos a la hoja Resumen
        thin = Side(border_style="thin", color="000000")
        borde = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # Identificar cambios de día
        fecha_actual = None
        color_dia_actual = None
        es_dia_par = False
        
        for i, row in enumerate(ws_resumen.iter_rows(), start=1):
            # Saltar encabezados
            if i == 1:
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    cell.border = borde
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                continue
            
            # Determinar si cambió el día
            fecha_celda = row[0].value  # Primera columna es fecha
            if fecha_celda != fecha_actual:
                fecha_actual = fecha_celda
                es_dia_par = not es_dia_par
                color_dia_actual = color_dia_par if es_dia_par else color_dia_impar
                fila_en_dia = 0  # Reiniciar contador de filas por día
            
            # Alternar tonos para filas del mismo día
            fila_en_dia += 1
            if fila_en_dia % 2 == 1:
                color_fila = color_dia_actual
            else:
                if es_dia_par:
                    color_fila = color_fila_alterna_oscuro
                else:
                    color_fila = color_fila_alterna_oscuro_verde
            
            # Aplicar estilo a toda la fila
            for cell in row:
                cell.fill = PatternFill(start_color=color_fila, end_color=color_fila, fill_type="solid")
                cell.border = borde
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_resumen.freeze_panes = "A2"
    
    # Mantener el estilo original para la hoja Detalle
    if 'Detalle' in wb.sheetnames:
        ws_detalle = wb['Detalle']
        colores_detalle = [
            "B7E1FC", "FFCCCC", "CCFFCC", "FFFFCC",
            "C6EFCE", "FFF2CC", "E4C6EF"
        ] * ((len(ws_detalle[1]) + 6) // 7)  # Paréntesis cerrado correctamente
        aplicar_estilos(ws_detalle, colores_detalle[:len(ws_detalle[1])])
    

    wb.save(path_excel)
    print(f"✅ Excel generado correctamente con nuevo diseño: {path_excel}")

def aplicar_estilos(ws, colores):
    """Función auxiliar para aplicar estilos básicos"""
    thin = Side(border_style="thin", color="000000")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.column <= len(colores):
                fill = PatternFill(
                    start_color=colores[cell.column-1],
                    end_color=colores[cell.column-1],
                    fill_type="solid"
                )
                cell.fill = fill
            
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if cell.row == 1:
                cell.font = Font(bold=True)
    
    ws.freeze_panes = "A2"

def guardar_excel(df_detalle, archivo):
    try:
        df_detalle = df_detalle.replace(r'^\s*$', pd.NA, regex=True)
        resumen_df = generar_hoja_resumen(df_detalle)

        archivo_existe = os.path.exists(archivo)

        if archivo_existe:
            wb = load_workbook(archivo)
            sheets_to_remove = [name for name in ['Detalle', 'Resumen'] if name in wb.sheetnames]
            for name in sheets_to_remove:
                wb.remove(wb[name])

            # Solo guardar si todavía hay hojas visibles
            if wb.sheetnames:
                wb.save(archivo)
            else:
                # Eliminar el archivo y recrearlo desde cero
                wb.close()
                os.remove(archivo)
            mode = 'a' if os.path.exists(archivo) else 'w'
        else:
            mode = 'w'


        with pd.ExcelWriter(archivo, engine='openpyxl', mode=mode) as writer:
            df_detalle.to_excel(writer, sheet_name='Detalle', index=False)
            resumen_df.to_excel(writer, sheet_name='Resumen', index=False)

        estilizar_excel(archivo)
        print(f"✅ Excel generado correctamente con nuevo diseño: {archivo}")

    except PermissionError:
        print(f"❌ Error: No se pudo guardar '{archivo}'. Cierre el archivo e intente nuevamente.")
        input("Presione Enter para continuar...")
        guardar_excel(df_detalle, archivo)
    except Exception as e:
        print(f"❌ Error inesperado: {str(e)}")
        raise

def generar_excel(desde_fecha: str, cantidad_dias: int, archivo_salida: str):
    fecha_inicio = datetime.strptime(desde_fecha, "%d-%m-%Y")
    columnas_totales = []
    bloques_por_dia = []
    max_filas = 0

    for dia in range(cantidad_dias):
        fecha_actual = fecha_inicio + timedelta(days=dia)
        fecha_str = fecha_actual.strftime("%d-%m-%Y")
        columnas = [fecha_str, "Cliente", "Cotización", "Segmento", "Tasa", "Ranking", "Venta Ponderada"]
        columnas_totales.extend(columnas)

        clientes_dia = generar_clientes_dia(fecha_actual)
        registros_dia = []
        
        for cliente in clientes_dia:
            cotizaciones = generar_cotizaciones_cliente(cliente)
            for cotizacion in cotizaciones:
                segmento = asignar_segmento(cliente, cotizacion)
                registros_dia.append([cliente, cotizacion, segmento, "", "", ""])  # 6 elementos + fecha = 7
        
        bloques_por_dia.append(registros_dia)
        max_filas = max(max_filas, len(registros_dia))

    filas_finales = []
    for fila_idx in range(max_filas):
        fila = []
        for dia_idx, bloque in enumerate(bloques_por_dia):
            fecha_str = (fecha_inicio + timedelta(days=dia_idx)).strftime("%d-%m-%Y")
            if fila_idx < len(bloque):
                # Añadir fecha + los 6 campos del registro
                fila.extend([fecha_str] + bloque[fila_idx])
            else:
                # Añadir fecha + 6 campos vacíos
                fila.extend([fecha_str] + ["", "", "", "", "", ""])
        filas_finales.append(fila)

    # Verificación de consistencia
    num_columnas_esperadas = cantidad_dias * 7
    num_columnas_reales = len(columnas_totales)
    if num_columnas_esperadas != num_columnas_reales:
        raise ValueError(f"Inconsistencia en columnas: esperadas {num_columnas_esperadas}, reales {num_columnas_reales}")

    df = pd.DataFrame(filas_finales, columns=columnas_totales)
    guardar_excel(df, archivo_salida)
    print(f"✅ Excel generado correctamente como: {archivo_salida}")

def procesar_rankings(df: pd.DataFrame):
    columnas = df.columns.tolist()
    dias_procesados = 0
    dias_omitidos = []

    for i in range(0, len(columnas), 7):  # Avanzamos de 7 en 7 columnas
        fecha_col = columnas[i]
        cliente_col = columnas[i+1]
        segmento_col = columnas[i+3]
        tasa_col = columnas[i+4]
        ranking_col = columnas[i+5]

        # Solo considerar filas donde hay cliente (ignorar filas vacías)
        filas_con_clientes = df[cliente_col].notna()
        
        if not filas_con_clientes.any():
            continue  # No hay clientes este día, saltar

        tasas = df.loc[filas_con_clientes, tasa_col]

        if tasas.isna().any():
            print(f"⚠️ Día {fecha_col} tiene tasas incompletas para algunos clientes. Se omite.")
            dias_omitidos.append(fecha_col)
            continue

        for fila in df[filas_con_clientes].index:
            if pd.notna(df.at[fila, ranking_col]):
                continue
            segmento = df.at[fila, segmento_col]
            tasa = df.at[fila, tasa_col]
            df.at[fila, ranking_col] = calcular_ranking(segmento, float(tasa))

        dias_procesados += 1

    return df, dias_procesados, dias_omitidos

def procesar_ventas_ponderadas(df: pd.DataFrame):
    columnas = df.columns.tolist()
    dias_calculados = 0

    for i in range(0, len(columnas), 7):
        # Validación de columnas mínimas
        if i + 6 >= len(columnas):
            continue
        
        cliente_col = columnas[i+1]
        segmento_col = columnas[i+3]
        ranking_col = columnas[i+5]
        venta_col = columnas[i+6]

        # Solo filas con cliente y ranking válido
        filas_validas = df[cliente_col].notna() & df[ranking_col].notna()

        for fila in df[filas_validas].index:
            segmento = df.at[fila, segmento_col]
            ranking = df.at[fila, ranking_col]
            try:
                venta = calcular_venta_ponderada(segmento, ranking)
                df.at[fila, venta_col] = round(float(venta), 2)
            except Exception as e:
                print(f"❌ Error en fila {fila}: {e}")
                df.at[fila, venta_col] = ""

        dias_calculados += 1

    return df, dias_calculados

def generar_resumen(df_detalle):
    resumen = []

    columnas = df_detalle.columns.tolist()
    for i in range(0, len(columnas), 7):
        if i + 6 >= len(columnas):
            continue  # Saltar si el bloque está incompleto

        fecha_col = columnas[i]
        segmento_col = columnas[i+3]
        ranking_col = columnas[i+5]

        filas_validas = df_detalle[segmento_col].notna() & df_detalle[ranking_col].notna()
        total = filas_validas.sum()
        if total == 0:
            continue

        conteo = df_detalle[filas_validas][ranking_col].value_counts().to_dict()
        fila_resumen = {
            "Fecha": df_detalle[fecha_col][0],
            "Total Clientes": total
        }

        for r in range(1, 6):
            fila_resumen[f"Ranking {r}"] = round(100 * conteo.get(r, 0) / total, 1)

        # Agrupamos ranking 6 o más
        otros = sum(v for k, v in conteo.items() if isinstance(k, (int, float)) and k >= 6)
        fila_resumen["Ranking 6 y más %"] = round(100 * otros / total, 1)

        resumen.append(fila_resumen)

    return pd.DataFrame(resumen)

def menu_repeticion(df, archivo):
    generar_resumen(df)  # Mostrar resumen antes del menú
    
    while True:
        print("\n¿Desea repetir alguna etapa?")
        print("1. Rellenar tasas y recalcular rankings")
        print("2. Modificar rankings manualmente y recalcular ventas ponderadas")
        print("3. Salir")
        opcion = input("Ingrese una opción (1/2/3): ")

        if opcion == '1':
            input("✏️ Rellene las tasas en el Excel y presione Enter para continuar...")
            df = pd.read_excel(archivo)
            df, _, _ = procesar_rankings(df)
            guardar_excel(df, archivo)
            generar_resumen(df)  # Mostrar resumen actualizado

        elif opcion == '2':
            input("✏️ Modifique manualmente los rankings y presione Enter para continuar...")
            df = pd.read_excel(archivo)
            df, _ = procesar_ventas_ponderadas(df)
            guardar_excel(df, archivo)
            generar_resumen(df)  # Mostrar resumen actualizado

        elif opcion == '3':
            print("👋 Finalizando...")
            break
        else:
            print("❌ Opción no válida. Intente de nuevo.")

def rellenar_tasas_en_excel(archivo):
    """Modifica directamente la hoja 'Detalle' del archivo Excel, rellenando tasas por segmento"""
    wb = load_workbook(archivo)
    
    if "Detalle" not in wb.sheetnames:
        print("❌ La hoja 'Detalle' no existe.")
        return
    
    ws = wb["Detalle"]
    columnas = [cell.value for cell in ws[1]]

    for i in range(0, len(columnas), 7):
        if i + 4 >= len(columnas):
            continue
        segmento_col = i + 4  # columna índice real (Excel es base 1)
        tasa_col = i + 5

        for fila in range(2, ws.max_row + 1):
            segmento = ws.cell(row=fila, column=segmento_col).value
            tasa = ws.cell(row=fila, column=tasa_col).value

            if segmento is not None and tasa in (None, ""):
                try:
                    segmento_int = int(segmento)
                    tasa_asignada = tasas_por_segmento.get(segmento_int, 0.5)
                    ws.cell(row=fila, column=tasa_col).value = tasa_asignada
                except:
                    continue

    wb.save(archivo)
    print("✅ Tasas automáticas asignadas en la hoja Detalle.")

def rellenar_ventas_ponderadas_en_excel(archivo):
    """Modifica directamente la hoja 'Detalle' del Excel, rellenando ventas ponderadas si hay ranking y segmento válidos"""
    from openpyxl import load_workbook

    wb = load_workbook(archivo)

    if "Detalle" not in wb.sheetnames:
        print("❌ La hoja 'Detalle' no existe.")
        return

    ws = wb["Detalle"]
    columnas = [cell.value for cell in ws[1]]

    for i in range(0, len(columnas), 7):
        if i + 6 >= len(columnas):
            continue

        segmento_col = i + 4 + 1  # Excel base 1
        ranking_col = i + 6
        venta_col = i + 7

        for fila in range(2, ws.max_row + 1):
            try:
                segmento = ws.cell(row=fila, column=segmento_col).value
                ranking = ws.cell(row=fila, column=ranking_col).value
                venta_actual = ws.cell(row=fila, column=venta_col).value

                if segmento is None or ranking is None or venta_actual not in (None, ""):
                    continue

                segmento = float(segmento)
                ranking = float(ranking)

                if ranking == 0:
                    print(f"⚠️ Ranking igual a cero en fila {fila}, se omite cálculo de venta.")
                    continue

                venta = calcular_venta_ponderada(segmento, ranking)
                ws.cell(row=fila, column=venta_col).value = round(float(venta), 2)

            except Exception as e:
                print(f"❌ Error en fila {fila}: {e}")
                continue

    wb.save(archivo)
    print("✅ Ventas ponderadas calculadas y asignadas directamente en hoja Detalle.")


# Ejemplo de uso:
if __name__ == "__main__":
    archivo = "datos_clientes.xlsx"

    # Configuración inicial
    tasas_por_segmento = {segmento: round(random.uniform(0.1, 1.0), 2) for segmento in range(1, 11)}
    
    # 1. Solicitar parámetros iniciales
    print("⚙️ CONFIGURACIÓN INICIAL")
    desde = input("📅 Ingrese la fecha de inicio (dd-mm-aaaa): ")
    while True:
        try:
            datetime.strptime(desde, "%d-%m-%Y")
            break
        except ValueError:
            desde = input("❌ Formato incorrecto. Use dd-mm-aaaa: ")

    while True:
        try:
            dias = int(input("📆 ¿Cuántos días desea simular?: "))
            if dias > 0:
                break
            print("❌ Ingrese un número positivo.")
        except ValueError:
            print("❌ Ingrese un número válido.")

    # 2. Generar el Excel base
    print("\n⏳ Generando archivo Excel inicial...")
    generar_excel(desde, dias, archivo)
    df = pd.read_excel(archivo)
    
    # 3. Función para rellenar tasas
    def rellenar_tasas(df, archivo):
        """Menú interactivo para rellenar tasas"""
        while True:
            print("\n🔧 OPCIONES PARA RELLENAR TASAS")
            print("1. Mostrar tasas predefinidas por segmento")
            print("2. Rellenar automáticamente basado en segmentos")
            print("3. Rellenar manualmente en Excel")
            print("4. Continuar a la siguiente etapa")
            
            opcion = input("Seleccione una opción (1/2/3/4): ")
            
            if opcion == '1':
                print("\n📊 Tasas predefinidas por segmento:")
                for seg, tasa in sorted(tasas_por_segmento.items()):
                    print(f"Segmento {seg}: {tasa}")
                continue
            
            elif opcion == '2':
                print("\n⏳ Asignando tasas automáticamente...")
                rellenar_tasas_en_excel(archivo)

                # Leer el archivo actualizado y regenerar hoja resumen
                df = pd.read_excel(archivo, sheet_name="Detalle")
                guardar_excel(df, archivo)  # Esto regenera hoja Resumen y re-aplica estilos

                print("✅ Tasas asignadas automáticamente y hoja Resumen actualizada.")

                revisar = input("¿Desea revisar/modificar las tasas manualmente? (s/n): ").lower()
                if revisar == 's':
                    input("✏️ Modifique las tasas en el Excel y presione Enter para continuar...")
                    df = pd.read_excel(archivo, sheet_name="Detalle")
                    guardar_excel(df, archivo)
                return df, True

                
            elif opcion == '3':
                input("✏️ Abra el archivo Excel, rellene las tasas manualmente y presione Enter para continuar...")
                df = pd.read_excel(archivo)
                guardar_excel(df, archivo)
                return df, True
                
            elif opcion == '4':
                return df, False
            else:
                print("❌ Opción no válida. Intente de nuevo.")

    # Mostrar menú de tasas hasta que estén completas o el usuario decida continuar
    while True:
        df, continuar = rellenar_tasas(df, archivo)
        if not continuar:
            break
        
        # Verificar si hay tasas faltantes (más robusto, maneja espacios y cadenas vacías)
        columnas = df.columns.tolist()
        tasas_faltantes = False

        for i in range(0, len(columnas), 7):
            if i + 4 >= len(columnas) or i + 3 >= len(columnas):
                continue
            segmento_col = columnas[i + 3]
            tasa_col = columnas[i + 4]

            # Filas con segmento válido (no NA, no vacío)
            filas_con_segmento = df[segmento_col].replace(r'^\s*$', pd.NA, regex=True).notna()

            # Para esas filas, limpiar tasa de espacios o cadenas vacías y revisar si está vacía
            tasas_limpias = df.loc[filas_con_segmento, tasa_col].replace(r'^\s*$', pd.NA, regex=True)

            if tasas_limpias.isna().any():
                faltantes = tasas_limpias.isna().sum()
                print(f"⚠️ La columna '{tasa_col}' tiene {faltantes} tasas faltantes para filas con segmento válido.")
                tasas_faltantes = True

        if tasas_faltantes:
            print("⚠️ Aún hay tasas faltantes en algunas cotizaciones con segmento asignado.")
        else:
            print("✅ Todas las tasas están correctamente asignadas en filas con segmento válido.")



        
        if not tasas_faltantes:
            break
        print("⚠️ Aún hay tasas faltantes en algunas cotizaciones.")

    # 4. Procesar rankings
    print("\n⏳ Calculando rankings...")
    while True:
        df, procesados, omitidos = procesar_rankings(df)
        guardar_excel(df, archivo)

        if omitidos:
            print(f"🔄 Días omitidos por tasas faltantes: {omitidos}")
            seguir = input("¿Desea rellenar las tasas faltantes y continuar? (s/n): ").lower()
            if seguir != 's':
                break
            df, _ = rellenar_tasas(df, archivo)
        else:
            break

    # 5. Calcular ventas ponderadas
    input("✏️ Puede modificar manualmente los rankings si desea. Presione Enter para continuar...")
    rellenar_ventas_ponderadas_en_excel(archivo)
    df = pd.read_excel(archivo)
    guardar_excel(df, archivo)

    # 6. Verificar ventas faltantes
    while True:
        vacios = []
        columnas = df.columns.tolist()
        
        for i in range(0, len(columnas), 7):
            if i + 6 >= len(columnas):  # 👈 PREVIENE EL IndexError
                print(f"⏭️ Saltando bloque incompleto desde columna {i}")
                continue

            fecha_col = columnas[i]
            cliente_col = columnas[i+1]
            venta_col = columnas[i+6]
            
            filas_con_clientes = df[cliente_col].notna()
            filas_con_ventas_faltantes = filas_con_clientes & df[venta_col].isna()
            
            if filas_con_ventas_faltantes.any():
                vacios.append(fecha_col)

        if vacios:
            print(f"📌 Días con ventas ponderadas faltantes para algunos clientes: {vacios}")
            seguir = input("¿Desea rellenar los rankings faltantes y continuar? (s/n): ").lower()
            if seguir == 's':
                df = pd.read_excel(archivo)
                df, _, _ = procesar_rankings(df)
                df, _ = procesar_ventas_ponderadas(df)
                guardar_excel(df, archivo)
            else:
                break
        else:
            break


    # 7. Menú de repetición y resumen final
    menu_repeticion(df, archivo)
    print("\n✅ Proceso completado. Archivo final guardado como:", archivo)